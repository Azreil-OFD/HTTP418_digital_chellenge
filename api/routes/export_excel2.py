from fastapi import APIRouter, Depends, HTTPException, Path
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
from fastapi.responses import StreamingResponse
from io import BytesIO
from api.database.db import get_session
from openpyxl import Workbook

router = APIRouter(tags=["export_excel"])

@router.get("/objects/export/{object_id}")
async def export_object_data(
    object_id: int = Path(..., description="ID of the object"),
    session: AsyncSession = Depends(get_session)
):
    try:
        object_type_query = text("SELECT type FROM objects WHERE id = :object_id")
        result = await session.execute(object_type_query, {'object_id': object_id})
        object_type_row = result.fetchone()
        if not object_type_row:
            raise HTTPException(status_code=404, detail="Object not found.")
        object_type = object_type_row[0]

        hierarchy_levels = ['mest', 'ngdu', 'cdng', 'kust', 'well']
        type_to_levels = {
            'mest': ['mest', 'ngdu', 'cdng', 'kust', 'well'],
            'ngdu': ['ngdu', 'cdng', 'kust', 'well'],
            'cdng': ['cdng', 'kust', 'well'],
            'kust': ['kust', 'well'],
            'well': ['well']
        }

        levels_to_process = type_to_levels.get(object_type)
        if not levels_to_process:
            raise HTTPException(status_code=400, detail="Invalid object type.")

        object_type_column = object_type  

        workbook = Workbook()
        workbook.remove(workbook.active) 

        for table_name in ['well_day_histories', 'well_day_plans']:
            for level in levels_to_process:
                if level == 'well' and object_type == 'well':
                    query = f"""
                        SELECT date, debit, ee_consume, expenses, pump_operating
                        FROM {table_name}
                        WHERE well = :object_id
                        ORDER BY date;
                    """
                    params = {'object_id': object_id}
                    columns = ['Date', 'Debit', 'EE Consume', 'Expenses', 'Pump Operating']
                    sheet_name = f"{object_type}_{object_id}_{table_name}_{level}"
                elif level == 'well':
                    query = f"""
                        SELECT WH.date, W.well, WH.debit, WH.ee_consume, WH.expenses, WH.pump_operating
                        FROM {table_name} WH
                        JOIN wells W ON WH.well = W.well
                        WHERE W.{object_type_column} = :object_id
                        ORDER BY WH.date, W.well;
                    """
                    params = {'object_id': object_id}
                    columns = ['Date', 'Well', 'Debit', 'EE Consume', 'Expenses', 'Pump Operating']
                    sheet_name = f"{object_type}_{object_id}_{table_name}_{level}"
                else:
                    query = f"""
                        SELECT WH.date, W.{level}, SUM(WH.debit) AS debit, SUM(WH.ee_consume) AS ee_consume,
                               SUM(WH.expenses) AS expenses, SUM(WH.pump_operating) AS pump_operating
                        FROM {table_name} WH
                        JOIN wells W ON WH.well = W.well
                        WHERE W.{object_type_column} = :object_id
                        GROUP BY WH.date, W.{level}
                        ORDER BY WH.date, W.{level};
                    """
                    params = {'object_id': object_id}
                    columns = ['Date', level.capitalize(), 'Debit', 'EE Consume', 'Expenses', 'Pump Operating']
                    sheet_name = f"{object_type}_{object_id}_{table_name}_{level}"

                data_result = await session.execute(text(query), params)
                data = data_result.fetchall()

                if data:
                    sheet_name = sheet_name[:31]
                    worksheet = workbook.create_sheet(title=sheet_name)

                    worksheet.append(columns)

                    for row in data:
                        worksheet.append(row)

        if not workbook.sheetnames:
            raise HTTPException(status_code=404, detail="No data found for the specified object.")

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        headers = {
            'Content-Disposition': f'attachment; filename="object_data_{object_id}.xlsx"'
        }
        return StreamingResponse(output, headers=headers)

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"An error occurred: {str(e)}")
